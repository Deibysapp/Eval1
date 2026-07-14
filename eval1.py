import streamlit as st
import pandas as pd
import streamlit.components.v1 as components
from openpyxl import load_workbook
from io import BytesIO

# Configuración de página
st.set_page_config(page_title="Evaluación Asesores GO", layout="wide")

# Lógica de almacenamiento local (No requiere instalar librerías extra)
def configurar_almacenamiento_local():
    js_code = """
    <script>
        // Guardar automáticamente al cambiar cualquier input
        window.parent.document.addEventListener('change', (e) => {
            if (e.target.id) {
                localStorage.setItem(e.target.id, e.target.value);
            }
        });
        // Recuperar datos al cargar la página
        window.onload = function() {
            setTimeout(() => {
                const inputs = window.parent.document.querySelectorAll('input, select');
                inputs.forEach(input => {
                    const savedValue = localStorage.getItem(input.id);
                    if (savedValue !== null) {
                        input.value = savedValue;
                        input.dispatchEvent(new Event('change'));
                    }
                });
            }, 1000);
        };
    </script>
    """
    components.html(js_code, height=0)

configurar_almacenamiento_local()

# Estructura de evaluación
estructura = {
    "Actitud de Ventas": [("Buena presencia y puntualidad", 9), ("Secuencia de visitas de acuerdo al rutero", 10), ("Asistencia a reuniones, auditorias, contacto 1 a 1", 11), ("Organización y Uso de material de trabajo", 12), ("Toma pedidos y cobranzas en el movil y sincroniza en los tiempos oportunos", 13), ("Informa las incideancias de la ruta al supervisor en los tiempos oportunos", 14)],
    "Pasos de la Visita Organizada": [("Preparación de la visita", 16), ("Contacto y acercamiento", 17), ("Chequeo de Inventario", 18), ("Identificación de oportunidades en el PDV", 19), ("Ejecución del producto en stock de inventario", 20), ("Asesoramiento y negociación del pedido", 21), ("Cobranza y cierre de la visita", 22)],
    "Dominio de la Venta": [("Gestiona todo el portafolio de productos", 24), ("Ofrece innovaciones o promociones", 25), ("Revisión y actualización de indicadores de precios en PDV", 26), ("Negociacion de espacios adicionales", 27)],
    "Material P.O.P y Planogramas": [("Presencia de material P.O.P vigente", 29), ("Evalua el cumplimiento del Planograma de Marcas", 30), ("Garantiza que los exhibidores no esten invadidos", 31), ("Evalua limpieza", 32)],
    "PROFUNDIDAD DE LINEAS": [("Ofrece Marcas Parmalat, PAVECA, Polar, Pepsi", 34), ("Ofrece Marcas Isola, Alfonzo Rivas", 35), ("Ofrece Marcas Alinieve, Incosa, Alesca", 36), ("Ofrece Marcas Miceven, Puig, Gaduca, Calven", 37)]
}

st.title("📝 Evaluación Asesores GO (Modo Offline)")

# Campos de entrada
fecha = st.date_input("Fecha", key="fecha")
vendedor = st.text_input("Vendedor", key="vendedor")
ruta = st.text_input("Ruta", key="ruta")
evaluador = st.text_input("Evaluador", key="evaluador")

# Generación de la interfaz
for seg, items in estructura.items():
    st.subheader(f"📍 {seg}")
    for item, fila in items:
        with st.expander(item):
            cols = st.columns(5)
            for i in range(5):
                cols[i].selectbox(f"C{i+1}", [0,1,2,3,4], key=f"{item}_{i}")

# Lógica de Exportación
def exportar_excel():
    wb = load_workbook('Eval1.xlsx')
    ws = wb.active
    ws['E4'] = str(fecha)
    ws['E6'] = vendedor
    ws['H6'] = ruta
    ws['E7'] = evaluador
    
    columnas = ['H', 'I', 'J', 'K', 'L']
    for seg, items in estructura.items():
        for item, fila in items:
            for i in range(5):
                ws[f"{columnas[i]}{fila}"] = st.session_state.get(f"{item}_{i}", 0)
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

if st.button("📥 Exportar a Excel"):
    st.download_button("Descargar Archivo", data=exportar_excel(), file_name="Reporte_Final.xlsx")

if st.button("🧹 Limpiar datos"):
    st.write("Refresca la página para limpiar los datos del teléfono.")

